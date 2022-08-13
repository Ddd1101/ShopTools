# coding=UTF-8
import xlrd
import xlwt
from xlutils.copy import copy
import datetime
from openpyxl import *
import xlsxwriter
import os

path = os.path.dirname(__file__)

price_path = path + '/price.xlsx'

shuadanSheet = object()



def write_excel_xlsx_by_xy(path,value,sheetName,x,y):
    print("# 写入文件")
    workbook = load_workbook(path)
    worksheet = workbook.get_sheet_by_name(sheetName)

    workbook.save(path)

def write_excel_xlsx_append(path,value,sheetName):
    print("# 写入文件")
    workbook = load_workbook(path)
    index = len(value)  # 获取需要写入数据的行数
    worksheet = workbook.get_sheet_by_name(sheetName)  # 获取工作簿中所有表格中的的第一个表格

    rows_old = worksheet.max_row  # 获取表格中已存在的数据的行数

    for i in range(0, index):
        worksheet.append(value[i])
    workbook.save(path)  # 保存工作簿
    print("xls格式表格【追加】 "+sheetName+" 写入数据成功！")

def write_excel_xls(path, sheet_name, value):
    index = len(value)  # 获取需要写入数据的行数
    workbook = xlwt.Workbook()  # 新建一个工作簿
    sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格
    for i in range(0, index):
        for j in range(0, len(value[i])):
            print(value[i][j])
            sheet.write(i, j, value[i][j])  # 像表格中写入数据（对应的行和列）
    workbook.save(path)  # 保存工作簿
    print("xls格式表格写入数据成功！")


def write_excel_xls_append(path, value):
    index = len(value)  # 获取需要写入数据的行数
    # print(index)
    workbook = xlrd.open_workbook(path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
    new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
    new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
    for i in range(0, index):
        # print(i)
        # print(value[i][1])
        for j in range(0, len(value[i])):
            new_worksheet.write(i + rows_old, j, value[i][j])  # 追加写入数据，注意是从i+rows_old行开始写入
    new_workbook.save(path)  # 保存工作簿
    print("xls格式表格【追加】写入数据成功！")

def GetWorkBook(path):
    workbook = xlrd.open_workbook(path)

    return workbook

def AppendWrite(_workbook,i,j,value):
    sheets = _workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = _workbook.sheet_by_name(sheets[3])  # 获取工作簿中所有表格中的的第一个表格
    rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
    new_workbook = copy(_workbook)  # 将xlrd对象拷贝转化为xlwt对象
    new_worksheet = new_workbook.get_sheet(3)
    new_worksheet.write(i + rows_old, j, value)

    return new_workbook

def read_excel_xls(path):
    workbook = xlrd.open_workbook(path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[3])  # 获取工作簿中所有表格中的的第一个表格
    for i in range(1, worksheet.nrows):
        for j in range(0, worksheet.ncols):
            if j == 0:
                date = xlrd.xldate_as_tuple(worksheet.cell(i, j).value, 0)
                print(datetime.datetime(*date).strftime('%Y-%m-%d'))
            else:
                print(worksheet.cell_value(i, j), "\t", end="")  # 逐行逐列读取数据
        print()

def GetPriceGrid():
    workbook = xlrd.open_workbook(price_path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    return worksheet

# def GetShuaDanSheet():
#     workbook = xlrd.open_workbook(price_path)  # 打开工作簿
#     sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
#     print(sheets[1])
#     worksheet = workbook.sheet_by_name(sheets[1])  # 获取工作簿中所有表格中的的第一个表格
#     return worksheet

def SetShuaDanSheetOrderIdSheet(workbookPath, sheetNum=0):
    workbook = xlrd.open_workbook(workbookPath)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    global shuadanSheet
    shuadanSheet = workbook.sheet_by_name(sheets[sheetNum])  # 获取工作簿中所有表格中的的第一个表格

def IsShuadan(orderId, col=0):
    # if type(shuadanSheet) is object:
    #     return False
    for t in range(1, shuadanSheet.nrows):
        # print(shuadanSheet.cell(t, col).value)
        # print(shuadanSheet.cell(t, col).value)
        if orderId == shuadanSheet.cell(t, col).value:
            # print(orderId == shuadanSheet.cell(t, col).value)
            return True
    return False

# book_name_xls = 'D:/workplace_pycharm/AliERP\AliERP/AliBackData/店铺1收支.xlsx'
#
# sheet_name_xls = 'xls格式测试表'
#
# value_title = [["姓名", "性别", "年龄", "城市", "职业"], ]
#
# value1 = [["张三", "男", "19", "杭州", "研发工程师"],
#           ["李四", "男", "22", "北京", "医生"],
#           ["王五", "女", "33", "珠海", "出租车司机"], ]
#
# value2 = [["Tom", "男", "21", "西安", "测试工程师"],
#           ["Jones", "女", "34", "上海", "产品经理"],
#           ["Cat", "女", "56", "上海", "教师"], ]
#
# write_excel_xls(book_name_xls, sheet_name_xls, value_title)
# write_excel_xls_append(book_name_xls, value1)
# write_excel_xls_append(book_name_xls, value2)
# read_excel_xls(book_name_xls)
